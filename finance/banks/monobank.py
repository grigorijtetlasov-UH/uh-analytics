"""
finance/banks/monobank.py
Адаптер Monobank personal API → financial.transactions.

Обмеження Monobank, враховані тут:
  • Rate-limit: 1 запит / 60 c на токен → спимо 61 c (глобально на всі токени).
  • Суми в копійках (÷100). currencyCode ISO-4217 numeric (980=UAH).
  • Виписка: макс. 31 день за запит, до 500 транзакцій (далі — пагінація назад).

Ідентичність ФОПа береться з API (clientId/name), НЕ з Моно.xlsx —
у xlsx колонка імені може бути константою і не розрізняє ФОПів.
"""
from __future__ import annotations

import logging
import re
import sys
import time
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Iterable, Optional

import requests
from openpyxl import load_workbook

from finance.banks.base import Account, BankAdapter, Transaction

API_BASE = "https://api.monobank.ua"
RATE_LIMIT_SEC = 61
MAX_WINDOW = timedelta(days=31)
PAGE_LIMIT = 500
CURRENCY_BY_CODE = {980: "UAH", 840: "USD", 978: "EUR"}


class MonobankAdapter(BankAdapter):
    bank_code = "mono"

    def __init__(self, creds_dir, logger=None, rate_limit_sec: int = RATE_LIMIT_SEC):
        super().__init__(creds_dir, logger)
        self.rate_limit_sec = rate_limit_sec
        self._creds: list[tuple[str, str]] = []        # [(label, token), ...]  label — лише для логів
        self._token_by_ext: dict[str, str] = {}        # clientId -> token (заповнює fetch_accounts)
        self._last_request_at = 0.0                     # monotonic; спільний на всі токени

    # ── credentials: потрібен ЛИШЕ стовпець токенів ────────────────
    def load_credentials(self) -> None:
        path = self.creds_dir / "Моно.xlsx"
        if not path.exists():
            raise FileNotFoundError(f"Не знайдено {path}")
        wb = load_workbook(path, read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        if not rows:
            raise ValueError("Моно.xlsx порожній")

        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        lbl_i = self._find_col(header, ["фіо", "піб", "фоп", "name", "ім"])
        tok_i = self._find_col(header, ["token", "токен", "ключ", "x-token"])

        creds = []
        for row in rows[1:]:
            if tok_i is not None:
                token = row[tok_i]
                label = row[lbl_i] if lbl_i is not None else None
            else:
                label, token = self._guess_label_token(row)
            if token:
                creds.append((str(label).strip() if label else "?", str(token).strip()))
        if not creds:
            raise ValueError("Не знайдено токенів у Моно.xlsx — перевір заголовки")
        self._creds = creds
        self.log.info("Завантажено %d Monobank токенів", len(creds))

    @staticmethod
    def _find_col(header, needles) -> Optional[int]:
        for i, h in enumerate(header):
            if any(n in h for n in needles):
                return i
        return None

    @staticmethod
    def _guess_label_token(row):
        label = token = None
        for cell in row:
            if cell is None:
                continue
            s = str(cell).strip()
            if not s:
                continue
            if " " not in s and len(s) >= 30 and re.fullmatch(r"[A-Za-z0-9_\-]+", s):
                token = s
            elif re.search(r"[А-Яа-яІіЇїЄєҐґ]", s):
                label = s
        return label, token

    # ── HTTP з глобальним rate-limit ───────────────────────────────
    def _request(self, token: str, path: str):
        wait = self.rate_limit_sec - (time.monotonic() - self._last_request_at)
        if wait > 0:
            self.log.debug("rate-limit: чекаю %.0f c", wait)
            time.sleep(wait)
        last = None
        for attempt in (1, 2):
            last = requests.get(API_BASE + path, headers={"X-Token": token}, timeout=30)
            self._last_request_at = time.monotonic()
            if last.status_code == 429:
                self.log.warning("429 від Monobank — чекаю %d c (спроба %d)",
                                 self.rate_limit_sec, attempt)
                time.sleep(self.rate_limit_sec)
                continue
            last.raise_for_status()
            return last.json()
        last.raise_for_status()
        return last.json()

    # ── accounts: ідентичність із API ──────────────────────────────
    def fetch_accounts(self) -> list[Account]:
        if not self._creds:
            self.load_credentials()
        out: list[Account] = []
        for label, token in self._creds:
            try:
                info = self._request(token, "/personal/client-info")
            except Exception as e:
                self.log.error("client-info FAILED (%s): %s", label, e)
                continue
            client_id = str(info.get("clientId") or label)
            name = info.get("name") or label
            self._token_by_ext[client_id] = token
            accs = info.get("accounts", [])
            for acc in accs:
                cur = CURRENCY_BY_CODE.get(acc.get("currencyCode"), str(acc.get("currencyCode")))
                out.append(Account(
                    external_id=client_id,            # унікальний clientId з API
                    fio=name,                         # ім'я з банку
                    iban=acc.get("iban", "") or "",
                    currency=cur,
                    balance=Decimal(acc.get("balance", 0)) / 100,
                    provider_account_id=acc.get("id"),
                    raw=acc,
                ))
            self.log.info("%s [%s]: %d рахунків", name, client_id, len(accs))
        return out

    # ── transactions ───────────────────────────────────────────────
    def fetch_transactions(self, account: Account,
                           date_from: date, date_to: date) -> Iterable[Transaction]:
        token = self._token_by_ext.get(account.external_id)
        if token is None:
            raise ValueError("Спершу виклич fetch_accounts() — мапа clientId→token порожня")
        acc_id = account.provider_account_id or "0"
        frm = datetime.combine(date_from, datetime.min.time(), tzinfo=timezone.utc)
        to = datetime.combine(date_to, datetime.max.time().replace(microsecond=0),
                              tzinfo=timezone.utc)
        window = frm
        while window < to:
            chunk_to = min(window + MAX_WINDOW, to)
            yield from self._fetch_window(token, acc_id, account, window, chunk_to)
            window = chunk_to

    def _fetch_window(self, token, acc_id, account, frm, to) -> Iterable[Transaction]:
        cur_to = to
        while True:
            f, t = int(frm.timestamp()), int(cur_to.timestamp())
            items = self._request(token, f"/personal/statement/{acc_id}/{f}/{t}")
            if not items:
                break
            for it in items:
                yield self._to_txn(it, account)
            if len(items) < PAGE_LIMIT:
                break
            oldest = min(it["time"] for it in items)
            cur_to = datetime.fromtimestamp(oldest - 1, tz=timezone.utc)
            if cur_to <= frm:
                break

    @staticmethod
    def _to_txn(it: dict, account: Account) -> Transaction:
        return Transaction(
            bank_txn_id=str(it["id"]),
            transaction_date=datetime.fromtimestamp(it["time"], tz=timezone.utc),
            amount=Decimal(it.get("amount", 0)) / 100,
            description=it.get("description") or "",
            counterpart_iban=it.get("counterIban") or None,
            counterpart_name=it.get("counterName") or None,
            currency=CURRENCY_BY_CODE.get(it.get("currencyCode"), account.currency),
            raw_data=it,
        )


# ── self-test ──────────────────────────────────────────────────────
def _self_test(live: bool):
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    creds_dir = Path("~/finance/credentials").expanduser()
    mono = MonobankAdapter(creds_dir)
    mono.load_credentials()
    print(f"\n✅ Токенів у Моно.xlsx: {len(mono._creds)}")
    for label, tok in mono._creds:
        print(f"   • token …{tok[-4:]}  (мітка: {label})")
    print("   (реальні імена/clientId ФОПів підтягнуться з API у --live)")
    if not live:
        print("\n(offline. Для перевірки API: python -m finance.banks.monobank --live)")
        return
    print("\n🌐 LIVE: client-info першого токена…")
    mono._creds = mono._creds[:1]
    accs = mono.fetch_accounts()
    for a in accs:
        print(f"   {a.fio} [{a.external_id}] | {a.iban} | {a.currency} | balance={a.balance}")
    if accs:
        print(f"\n🌐 Виписка за 3 дні (чекай ~61 c rate-limit)…")
        end, start = date.today(), date.today() - timedelta(days=3)
        n = 0
        for tx in mono.fetch_transactions(accs[0], start, end):
            n += 1
            if n <= 5:
                print(f"   {tx.transaction_date:%Y-%m-%d %H:%M} | {tx.amount:>10} | "
                      f"{tx.direction} | {tx.description[:30]}")
        print(f"   разом транзакцій: {n}")


if __name__ == "__main__":
    _self_test(live="--live" in sys.argv)
